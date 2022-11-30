###################################################
#                                                 #
#   Rocket Simulation                             #
#                                                 #
#   Taking into account air resistance            #
#   varying mass, and thrust                      #
#                                                 #
#                                                 #
###################################################

    ##  Importing libraries and spreadsheet ##
import thrustCurveAnalysis_UPDATED as tca
import matplotlib.pyplot as plt
import openpyxl
import PySimpleGUI as sg
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib
from openpyxl import load_workbook
from os.path import exists

wb = load_workbook("Mk1 Data Sheet.xlsx")
Engines = wb["Engines"]
Dims = wb["Volume Calculations"]
tcd = load_workbook("thrustCurveData.xlsx")
tcdSheet = tcd["Sheet1"]

saveData = True

if exists("coconut.jpg"):

  # choice=input("Which engine do you want to use? [1] D12  [2] F15  ")
  #pdfChoice=input("Do you want to use NAR data, or your own? [1] NAR  [2] Own  ")

  ## TODO: Add a way to choose between NAR data and own data ##

  # event, values = window.read()
  
  def calcuateSim(choice):

    numberOfEngines = Engines['D2'].value

    if choice == "D12":

        ##  Defining variables ##
      initialMass = (Engines['B11'].value+Engines['B12'].value)*numberOfEngines + Engines['B22'].value + Engines['B21'].value  # kg
      propellantMass = Engines['B12'].value*numberOfEngines  # kg
      burnRate = propellantMass / (Engines['B10'].value)  # burnRate in kg/s
      timeLimit = 12  # seconds
      timeStep = .001  # seconds
      angle = 14.04952  # Angle between edge and center of cone in degrees

    elif choice == "F15":
  
          ##  Defining variables ##
      initialMass = (Engines['G11'].value+Engines['G12'].value)*numberOfEngines + Engines['G22'].value + Engines['G21'].value  # kg
      propellantMass = Engines['G12'].value*numberOfEngines  # kg
      burnRate = propellantMass / (Engines['G10'].value)  # burnRate in kg/s
      timeLimit = 12  # seconds
      timeStep = .001  # seconds
      angle = 14.04952  # Angle between edge and center of cone in degrees

    crossSectionalArea = 3.14*(0.305/2)**2  # m^2
    coeffDrag = 0.0112*angle+0.162  # drag coefficient

    ##   Initializing Arrays   ##
    time = []
    Thrust = []
    accel = []
    Vel = []
    height = []

  ##   Zeroing Variables   ##
    burnedMass = 0
    velocity = 0
    distance = 0
    maxHeight = 0
    maxVel = 0
    Drag = 0
    density = 0
    temperature = 0
    pressure = 0
    thrust = 0

    thrustCurve=[]

    thrustCurve = tca.main(choice)


    #### Main Loop ####
    for i in range(0, int(timeLimit/timeStep)):
      if burnedMass < propellantMass:
          currentMass = initialMass-burnedMass

        ### Calculating Thrust from curve ###
      for j in thrustCurve:
        if j[0] >= i*timeStep:
          thrust = (j[1]*(i*timeStep)+j[2])*numberOfEngines
          break
        else:
          thrust = 0

        ######  This is where the drag is calculated  ######
      if distance < 11019.13:
          # (Assuming linear density change under 1km)
          density = 1.225-0.000113*distance
      Drag = coeffDrag*(density*velocity**2)/2*crossSectionalArea  # Newtons


        ######  This is where the kinematics are calculated  ######
      acceleration = (thrust-Drag)/currentMass-9.8
      velocity = velocity+acceleration*timeStep
      distance = distance+velocity*timeStep
      burnedMass = burnedMass+burnRate*(1*timeStep)

      ######  This is where the data is stored  ######
      time.append(i)
      height.append(distance)
      Thrust.append(thrust)
      Vel.append(velocity)
      accel.append(acceleration)

    maxVel = max(Vel)
    maxHeight = max(height)
    
    ##### Plotting #####
    fig, axis = plt.subplots(2,2)
    fig.suptitle('Max Height: '+str(round(maxHeight, 3))+'m Max Velocity: '+str(round(maxVel, 3))+'m/s')
    
    axis[0,0].plot(time, height)
    axis[0,0].set_title("Height vs Time")
    axis[0,1].plot(time, Vel)
    axis[0,1].set_title("Velocity vs Time")
    axis[1,0].plot(time, Thrust)
    axis[1,0].set_title("Thrust vs Time")
    axis[1,1].plot(time, accel)
    axis[1,1].set_title("Acceleration vs Time")
    
    if saveData == False:
      tcdSheet.delete_cols(1, 1000)
      tcdSheet.delete_rows(1, 1000)
      tcd.save("thrustCurveData.xlsx")
    
    return maxHeight, maxVel, time, height, Vel, Thrust, accel, fig
